import json
import logging
import os
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent))

import smtplib
from email.mime.text import MIMEText

from config.cities import CITIES, START_CITY
from config.verticals import VERTICALS
from enrichment.email import enrich_business_email
from output.database import get_lead_count, insert_lead
from scraper.browser import close_browser, launch_browser, check_proxy_health
from scraper.maps import get_business_details, search_google_maps
from scraper.reviews import process_reviews, scrape_reviews
from signals.qualify import format_phone_e164, qualify_business

from dotenv import load_dotenv
load_dotenv()

TEST_MODE       = os.getenv("TEST_MODE", "true").lower() == "true"
BATCH_TEST_MODE = os.getenv("BATCH_TEST_MODE", "false").lower() == "true"
MAX_BUSINESSES  = int(os.getenv("MAX_BUSINESSES", "0")) or None

# ── Logging ────────────────────────────────────────────────────────────────────
Path("logs").mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("logs/scraper.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ── Email alerts ──────────────────────────────────────────────────────────────
def send_alert(subject, body):
    """Send an email alert to all recipients. Silently skips if credentials not configured."""
    sender   = os.getenv("ALERT_EMAIL_FROM", "").strip()
    password = os.getenv("ALERT_EMAIL_PASSWORD", "").replace(" ", "")
    recipients = [r.strip() for r in os.getenv("ALERT_EMAIL_TO", "").split(",") if r.strip()]

    if not all([sender, password, recipients]):
        return

    try:
        msg = MIMEText(body)
        msg["Subject"] = f"[Altrium Scraper] {subject}"
        msg["From"]    = sender
        msg["To"]      = ", ".join(recipients)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(sender, password)
            for recipient in recipients:
                msg.replace_header("To", recipient)
                smtp.sendmail(sender, recipient, msg.as_string())
        log.info(f"Alert sent to {len(recipients)} recipients: {subject}")
    except Exception as e:
        log.warning(f"Could not send alert email: {e}")


# ── Clean business name ───────────────────────────────────────────────────────
def clean_business_name(name):
    """
    Strips descriptive suffixes from Google Maps business names.
    Removes everything after |, (, or [ and after space-dash-space.
    Preserves hyphens within names (e.g. Roto-Rooter).
    """
    if not name:
        return name
    for char in ['|', '(', '[']:
        if char in name:
            name = name[:name.index(char)]
    # Only strip on space-dash-space, not bare hyphens
    name = re.split(r'\s+[-–]\s+', name)[0]
    return name.strip().rstrip(',').strip()


# ── Review snippet ────────────────────────────────────────────────────────────
def create_review_snippet(text, max_chars=200):
    """
    Trims review text to max_chars for use in email templates.
    Cuts at a word boundary and appends '...' if truncated.
    """
    if not text:
        return ""
    # Collapse whitespace and newlines into single spaces
    clean = " ".join(text.split())
    if len(clean) <= max_chars:
        return clean
    # Cut at last space before max_chars so we never split a word
    trimmed = clean[:max_chars].rsplit(" ", 1)[0].rstrip(".,;:!?")
    return trimmed + "..."


# ── Excel export ──────────────────────────────────────────────────────────────
EXCEL_COLUMNS = [
    "business_name", "clean_name", "phone", "all_phones", "business_email",
    "all_emails", "email_provider", "city", "state", "vertical",
    "qualifying_signal", "star_rating", "total_reviews", "bad_review_date",
    "bad_review_stars", "bad_review_text", "review_snippet", "bad_review_author",
    "days_since_last_response", "maps_url",
]

HEADERS = [
    "Business Name", "Clean Name", "Phone", "All Phones", "Email",
    "All Emails", "Email Provider", "City", "State", "Vertical",
    "Signal", "Rating", "Reviews", "Bad Review Date",
    "Bad Stars", "Bad Review Text", "Review Snippet", "Bad Review Author",
    "Days Since Response", "Maps URL",
]


def export_to_excel(leads, path="logs/leads.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, lead in enumerate(leads, start=2):
        for col, key in enumerate(EXCEL_COLUMNS, start=1):
            value = lead.get(key)
            if key == "days_since_last_response" and value is None:
                value = "Never"
            ws.cell(row=row, column=col, value=value)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(path)
    log.info(f"  Excel saved → {path}")
    return path


# ── Core search ────────────────────────────────────────────────────────────────
def run_search(vertical, city, state, use_proxy=False, max_businesses=None):
    summary = {"searched": 0, "qualified": 0, "inserted": 0, "skipped_dupe": 0}
    summary["leads"] = []

    listings = []
    browser, page = None, None
    for attempt in range(3):
        if browser:
            close_browser(browser)
        browser, page = launch_browser(use_proxy=use_proxy)
        listings = search_google_maps(page, vertical, city, state)
        if listings:
            break
        if attempt < 2:
            log.info(f"  No listings — retrying with fresh proxy IP ({attempt + 1}/2)")

    try:
        summary["searched"] = len(listings)
        log.info(f"  Found {len(listings)} listings")
        for _b in listings[:3]:
            log.info(f"  SAMPLE: {_b.get('business_name')} | ★{_b.get('star_rating')} | {_b.get('total_reviews')} reviews")

        candidates = [
            b for b in listings
            if b.get("star_rating") is not None
            and b.get("total_reviews") is not None
            and 3.0 <= b["star_rating"] <= 4.9
            and b["total_reviews"] >= 15
        ]
        log.info(f"  {len(candidates)} pass base filter (3.0–4.9 ★, 15+ reviews)")

        if max_businesses:
            candidates = candidates[:max_businesses]
            log.info(f"  Capped at {max_businesses} businesses for this run")

        for biz in candidates:
            try:
                details = get_business_details(page, biz["maps_url"])

                if not details.get("phone"):
                    log.info(f"  ↷ Skipped {biz['business_name']} — no phone number found")
                    continue

                phone = format_phone_e164(details["phone"])

                raw_reviews = scrape_reviews(page, biz["maps_url"])
                reviews = process_reviews(raw_reviews)

                signal = qualify_business(reviews, {**biz, **details})
                if signal is None:
                    log.info(f"  ↷ Skipped {biz['business_name']} — no signal fired")
                    continue

                summary["qualified"] += 1
                log.info(
                    f"  ✓ Signal {signal['qualifying_signal'].upper()} — "
                    f"{biz['business_name']} ({biz['star_rating']}★, "
                    f"{biz['total_reviews']} reviews)"
                )

                biz_enriched = {**biz, **details, "phone": phone}
                enrich_business_email(biz_enriched)

                lead = {
                    "business_name":          biz["business_name"],
                    "clean_name":             clean_business_name(biz["business_name"]),
                    "phone":                  biz_enriched.get("phone") or phone,
                    "business_email":         biz_enriched.get("business_email"),
                    "all_emails":             biz_enriched.get("all_emails"),
                    "all_phones":             biz_enriched.get("all_phones"),
                    "email_provider":         biz_enriched.get("email_provider", "Unknown"),
                    "city":                   details.get("city") or city,
                    "state":                  details.get("state") or state,
                    "vertical":               vertical,
                    "qualifying_signal":      signal["qualifying_signal"],
                    "star_rating":            biz["star_rating"],
                    "total_reviews":          biz["total_reviews"],
                    "bad_review_date":        signal["bad_review_date"],
                    "bad_review_stars":       signal["bad_review_stars"],
                    "bad_review_text":        signal["bad_review_text"],
                    "review_snippet":         create_review_snippet(signal["bad_review_text"]),
                    "bad_review_author":      signal["bad_review_author"],
                    "days_since_last_response": signal["days_since_last_response"],
                    "maps_url":               biz["maps_url"],
                }

                inserted = insert_lead(lead)
                if inserted:
                    summary["inserted"] += 1
                    summary["leads"].append(lead)
                    log.info(f"    → Saved to Supabase: {phone}")
                else:
                    summary["skipped_dupe"] += 1
                    log.info(f"    → Duplicate skipped: {phone}")

            except Exception as e:
                log.error(f"  Error on {biz.get('business_name')}: {e}")
                continue

    finally:
        close_browser(browser)

    return summary


# ── Full pipeline ──────────────────────────────────────────────────────────────
def run_full_pipeline(verticals=None, cities=None):
    verticals = verticals or VERTICALS
    cities    = cities    or CITIES

    checkpoint_path = Path("logs/checkpoint.json")
    start_c, start_v = START_CITY, 0

    if checkpoint_path.exists():
        with open(checkpoint_path) as f:
            cp = json.load(f)
        start_c = cp.get("city_index", 0)
        start_v = cp.get("vertical_index", 0)
        log.info(f"Resuming from city {start_c}, vertical {start_v}")

    total_inserted = 0

    for ci, city_data in enumerate(cities):
        if ci < start_c:
            continue

        city, state = city_data["city"], city_data["state"]

        # Check proxy is alive before starting each new city
        if not check_proxy_health():
            msg = (
                f"Proxy health check failed before {city}, {state} "
                f"(city {ci+1}/{len(cities)}).\n\n"
                f"Checkpoint saved — top up your Decodo plan and restart "
                f"to resume from where it stopped."
            )
            log.error("━" * 60)
            log.error("  PROXY EXHAUSTED — scraper paused")
            log.error(f"  Stopped at: {city}, {state} (city {ci+1})")
            log.error("  Restart after topping up Decodo to resume")
            log.error("━" * 60)
            send_alert("Proxy exhausted — scraper paused", msg)
            sys.exit(1)

        for vi, vertical in enumerate(verticals):
            if ci == start_c and vi < start_v:
                continue

            log.info(f"[City {ci+1}/{len(cities)}] {city}, {state} — [{vi+1}/{len(verticals)}] {vertical}")
            summary = run_search(vertical, city, state, use_proxy=True)
            total_inserted += summary["inserted"]

            log.info(
                f"  → searched={summary['searched']}  qualified={summary['qualified']}"
                f"  inserted={summary['inserted']}  dupes={summary['skipped_dupe']}"
            )

            with open(checkpoint_path, "w") as f:
                json.dump({"city_index": ci, "vertical_index": vi + 1}, f)

            time.sleep(random.uniform(5, 15))

        # All verticals done for this city — advance city index, reset vertical
        with open(checkpoint_path, "w") as f:
            json.dump({"city_index": ci + 1, "vertical_index": 0}, f)

    log.info(f"Pipeline complete — total leads inserted: {total_inserted}")
    return total_inserted


# ── Batch test ─────────────────────────────────────────────────────────────────
def run_batch_test(city="Houston", state="TX"):
    """
    Runs all verticals for a single city with proxy.
    Designed to validate the full pipeline and collect 100+ real leads
    before deploying to VPS.
    """
    log.info("━" * 60)
    log.info(f"  ALTRIUM SCRAPER — BATCH TEST")
    proxy_on = os.getenv("USE_PROXY", "false").lower() == "true"
    log.info(f"  City: {city}, {state}  |  Verticals: {len(VERTICALS)}  |  Proxy: {'ON' if proxy_on else 'OFF'}")
    log.info("━" * 60)

    all_leads = []
    total_inserted = 0
    total_qualified = 0

    for vi, vertical in enumerate(VERTICALS, 1):
        log.info(f"[{vi}/{len(VERTICALS)}] {vertical} in {city}, {state}")
        try:
            summary = run_search(vertical, city, state, use_proxy=os.getenv("USE_PROXY", "false").lower() == "true", max_businesses=MAX_BUSINESSES)
        except Exception as e:
            log.error(f"  ✗ Skipping {vertical} — {e}")
            time.sleep(random.uniform(5, 15))
            continue

        all_leads.extend(summary["leads"])
        total_inserted += summary["inserted"]
        total_qualified += summary["qualified"]

        log.info(
            f"  → searched={summary['searched']}  qualified={summary['qualified']}"
            f"  inserted={summary['inserted']}  dupes={summary['skipped_dupe']}"
            f"  |  running total: {total_inserted} leads"
        )

        time.sleep(random.uniform(5, 15))

    db_count = get_lead_count()

    log.info("━" * 60)
    log.info(f"  Verticals run:       {len(VERTICALS)}")
    log.info(f"  Signals fired:       {total_qualified}")
    log.info(f"  Leads inserted:      {total_inserted}")
    log.info(f"  Total leads in DB:   {db_count}")
    log.info("━" * 60)

    if all_leads:
        export_to_excel(all_leads, "logs/batch_test_leads.xlsx")

    with open("logs/batch_test_summary.json", "w") as f:
        json.dump({
            "run_at":            datetime.now().isoformat(),
            "city":              f"{city}, {state}",
            "verticals_run":     len(VERTICALS),
            "leads_inserted":    total_inserted,
            "total_leads_in_db": db_count,
        }, f, indent=2)

    if total_inserted >= 100:
        log.info(f"  ✓ TARGET MET — {total_inserted} leads collected. Ready for VPS.")
    else:
        log.info(f"  ⚠ Only {total_inserted} leads — consider running a second city.")

    return total_inserted


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if BATCH_TEST_MODE:
        run_batch_test()
    elif TEST_MODE:
        log.info("━" * 60)
        log.info("  ALTRIUM SCRAPER — DEMO RUN")
        log.info("  Search: plumber in Houston, TX")
        log.info("━" * 60)

        summary = run_search("plumber", "Houston", "TX",use_proxy=True, max_businesses=MAX_BUSINESSES)
        count   = get_lead_count()

        with open("logs/test_run.json", "w") as f:
            json.dump({
                "run_at":             datetime.now().isoformat(),
                "search":             "plumber in Houston, TX",
                "summary":            summary,
                "total_leads_in_db":  count,
            }, f, indent=2)

        log.info("━" * 60)
        log.info(f"  Businesses found:     {summary['searched']}")
        log.info(f"  Passed base filter:   {summary['searched']} → signals checked")
        log.info(f"  Signals fired:        {summary['qualified']}")
        log.info(f"  Leads inserted:       {summary['inserted']}")
        log.info(f"  Duplicates skipped:   {summary['skipped_dupe']}")
        log.info(f"  Total leads in DB:    {count}")
        log.info("━" * 60)

        if summary["leads"]:
            export_to_excel(summary["leads"], "logs/leads.xlsx")

        if summary["qualified"] > 0:
            log.info("  ✓ ALL STAGES COMPLETE — SCRAPER WORKING")
        else:
            log.info("  No signals fired — all businesses had clean review profiles")

    else:
        run_full_pipeline()
